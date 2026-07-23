#!/usr/bin/env python3
"""
Migration V3 de pp1_collection.db (PP1/Photomaton)
- Ajoute artworks.thumbnail_url (NULL par défaut)
- Normalise artworks.media_url : minuscules, [_ .] -> '-', seuls [a-z0-9-] conservés
  (les extensions sont ajoutées ensuite par sync_images_local.py, exécuté sur la
   machine qui possède le dossier d'images)
- Crée le référentiel `emotions` (emotion, name_fr, name_nl, name_en)
- Ajoute artworks.keywords_fr / keywords_nl / keywords_en
- Applique les retraits validés par le musée (relecture juillet 2026) :
  5 témoignages supprimés, type 28 « Printed image » fusionné dans le 21
  « Print - Image » (voir section 6 en fin de script).
- Ré-importe TOUTES les traductions corrigées depuis review_translations.xlsx :
  descriptions, questions, types d'objets, transcriptions (translation_*),
  explications (explanation_*), témoignages enregistrés (content_*)
Usage: python migrate_v3.py <db_path> <review_translations.xlsx>
"""
import re
import sys
import sqlite3
from openpyxl import load_workbook

DB = sys.argv[1] if len(sys.argv) > 1 else 'pp1_collection.db'
XLSX = sys.argv[2] if len(sys.argv) > 2 else 'review_translations.xlsx'


# Correspondance (fichier_source, question) -> id de l'interview dans `testimonies`
# (ids historiques préservés lors de la fusion V8 ; interviews = visitor_id IS NULL).
# Figée car les colonnes speaker/source_file ont été supprimées de la table (V3.1).
# Stable : l'ordre d'insertion de pp1_to_sqlite.py (= ordre du CSV) produit ces ids.
RECORDED_IDS = {
    'André_Kortrijk': {1: 101, 5: 121, 7: 130, 9: 140, 11: 149},
    'Anne-Lise_Mons': {4: 27, 5: 35, 6: 44, 7: 52, 12: 92},
    'Benoit_Mons': {3: 18, 4: 28, 6: 45, 10: 76, 11: 84},
    'Bernard_Kortrijk': {1: 102, 6: 125, 9: 141, 12: 153},
    'Charlotte_Lens': {1: 1, 6: 46, 7: 53, 8: 61, 9: 68},
    'Claire_Kortrijk': {1: 103, 2: 106, 6: 126, 10: 144},
    'Cécile_Mons': {4: 29, 5: 36, 7: 54, 9: 69, 12: 93},
    'Els_Kortrijk': {2: 107, 3: 113, 4: 116, 7: 132, 8: 135, 9: 142, 10: 145},
    'Fiona_Lens': {2: 9, 5: 37, 8: 62, 9: 70, 12: 94},
    'Godelieve_Kortrijk': {8: 136, 12: 154},
    'Greta_Kortrijk': {1: 104, 4: 117, 5: 122, 8: 137, 11: 150},
    'Hilda_Kortrijk': {1: 105, 3: 115, 4: 118, 8: 138, 10: 146},
    'Isa_Mons': {1: 2, 2: 10, 7: 55, 10: 77, 11: 85},
    'Jacques_Kortrijk': {2: 109, 4: 119, 7: 133, 10: 147, 11: 151},
    'Johanna_Mons': {1: 3, 5: 38, 6: 47, 7: 56, 10: 78},
    'Manuela_Mons': {1: 4, 2: 11, 7: 57, 8: 63, 10: 79},
    'Marion_Lens': {2: 12, 4: 30, 5: 39, 7: 58, 10: 80},
    'Moussa_Lens': {1: 5, 3: 19, 6: 48, 8: 64, 11: 86},
    'Muriel_Mons': {2: 13, 3: 20, 8: 65, 9: 71, 11: 87},
    'Nora_Mons': {3: 21, 4: 31, 6: 49, 9: 72, 12: 95},
    'Océane_Mons': {2: 14, 3: 22, 5: 40, 8: 66, 12: 96},
    'Olivier_Lens': {3: 23, 5: 41, 6: 50, 9: 73, 12: 97},
    'Persoon8_Kortrijk': {2: 110, 5: 123, 6: 128, 9: 143, 12: 155},
    'Sarah_Lens': {1: 6, 4: 32, 7: 59, 10: 81, 11: 88},
    'Scott_Lens': {2: 15, 6: 51, 10: 82, 11: 89, 12: 98},
    'Thierry_Lens': {1: 7, 3: 24, 7: 60, 10: 83, 12: 99},
    'Thomas_Lens': {2: 16, 4: 33, 8: 67, 9: 74, 11: 90},
    'Vincent_Lens': {1: 8, 2: 17, 3: 25, 4: 34, 5: 42},
    'Virginie_Mons': {3: 26, 5: 43, 9: 75, 11: 91, 12: 100},
    'Wim_Kortrijk': {2: 111, 4: 120, 5: 124, 6: 129, 7: 134, 8: 139, 10: 148, 11: 152},
}

IMG_EXTS = ('.jpg', '.jpeg', '.png', '.webp', '.gif', '.tif', '.tiff', '.bmp', '.avif')


def normalize_media_name(name: str) -> str:
    """IMadeYou_01_053.JPG -> imadeyou-01-053.jpg
    Préserve le point de l'extension (fix V5 intégré : l'ancienne version le
    transformait en '-' au re-run, cassant les media_url déjà porteurs
    d'extensions)."""
    s = name.strip().lower()
    ext = ''
    for e in IMG_EXTS:
        if s.endswith(e):
            s, ext = s[:-len(e)], e
            break
    s = re.sub(r'[_.\s]+', '-', s)      # _ . espaces -> -  (stem uniquement)
    s = re.sub(r'[^a-z0-9-]', '', s)     # tout le reste supprimé (sauf -)
    s = re.sub(r'-{2,}', '-', s).strip('-')
    return s + ext


def normalize_media_url(value):
    if not value:
        return value
    parts = [p.strip() for p in str(value).split(',')]
    parts = [normalize_media_name(p) for p in parts if p.strip()]
    return ', '.join(parts)


def main():
    con = sqlite3.connect(DB)
    cur = con.cursor()
    wb = load_workbook(XLSX, data_only=True)
    report = []

    def col_exists(table, col):
        return col in [r[1] for r in cur.execute(f'PRAGMA table_info({table})')]

    # ── 1. thumbnail_url ────────────────────────────────────────────────
    if not col_exists('artworks', 'thumbnail_url'):
        cur.execute('ALTER TABLE artworks ADD COLUMN thumbnail_url TEXT')
        report.append('artworks.thumbnail_url ajoutée (NULL)')
    else:
        report.append('artworks.thumbnail_url déjà présente')

    # ── 2. normalisation media_url ──────────────────────────────────────
    changed = 0
    for aid, mu in cur.execute('SELECT id, media_url FROM artworks WHERE media_url IS NOT NULL').fetchall():
        norm = normalize_media_url(mu)
        if norm != mu:
            cur.execute('UPDATE artworks SET media_url=? WHERE id=?', (norm, aid))
            changed += 1
    report.append(f'media_url normalisés : {changed} lignes modifiées')

    # ── 3. référentiel emotions ─────────────────────────────────────────
    cur.execute('''CREATE TABLE IF NOT EXISTS emotions (
        emotion TEXT PRIMARY KEY,
        name_fr TEXT NOT NULL,
        name_nl TEXT NOT NULL,
        name_en TEXT NOT NULL)''')
    ws = wb['Emotions']
    n = 0
    for src, fr, nl, en in ws.iter_rows(min_row=2, values_only=True):
        if not src:
            continue
        key = str(src).strip().lower()
        cur.execute('INSERT OR REPLACE INTO emotions VALUES (?,?,?,?)',
                    (key, str(fr).strip(), str(nl).strip(), str(en).strip()))
        n += 1
    report.append(f'référentiel emotions : {n} entrées (feuille)')
    # Typos dans artwork_emotions -> normalisation des données
    fixed = cur.execute("UPDATE artwork_emotions SET emotion='submission' WHERE lower(trim(emotion))='submision'").rowcount
    fixed += cur.execute("UPDATE artwork_emotions SET emotion='disapproval' WHERE lower(trim(emotion))='dissaproval'").rowcount
    report.append(f'typos corrigées dans artwork_emotions (submision/dissaproval) : {fixed} lignes')
    # Complément Plutchik pour les émotions utilisées mais absentes de la feuille (A VALIDER)
    COMPLEMENT = {
        'aggressiveness': ('Agressivité', 'Agressiviteit', 'Aggressiveness'),
        'amazement':      ('Stupéfaction', 'Verbazing', 'Amazement'),
        'anger':          ('Colère', 'Woede', 'Anger'),
        'annoyance':      ('Agacement', 'Ergernis', 'Annoyance'),
        'contempt':       ('Mépris', 'Minachting', 'Contempt'),
        'disapproval':    ('Désapprobation', 'Afkeuring', 'Disapproval'),
        'interest':       ('Intérêt', 'Interesse', 'Interest'),
        'pensiveness':    ('Rêverie', 'Bedachtzaamheid', 'Pensiveness'),
        'remorse':        ('Remords', 'Wroeging', 'Remorse'),
        'sadness':        ('Tristesse', 'Verdriet', 'Sadness'),
        'surprise':       ('Surprise', 'Verrassing', 'Surprise'),
    }
    added = 0
    for key, (fr, nl, en) in COMPLEMENT.items():
        added += cur.execute('INSERT OR IGNORE INTO emotions VALUES (?,?,?,?)', (key, fr, nl, en)).rowcount
    report.append(f'référentiel emotions : {added} entrées complétées (Plutchik, à valider)')
    # contrôle de couverture : chaque emotion utilisée doit exister dans le référentiel
    missing = cur.execute('''SELECT DISTINCT lower(trim(ae.emotion)) FROM artwork_emotions ae
        LEFT JOIN emotions e ON e.emotion = lower(trim(ae.emotion))
        WHERE e.emotion IS NULL''').fetchall()
    report.append(f'emotions utilisées sans traduction : {[m[0] for m in missing] or "aucune"}')

    # ── 4. keywords i18n ────────────────────────────────────────────────
    for c in ('keywords_fr', 'keywords_nl', 'keywords_en'):
        if not col_exists('artworks', c):
            cur.execute(f'ALTER TABLE artworks ADD COLUMN {c} TEXT')
    ws = wb['Description - Key words']
    n = miss = 0
    for i, _src, fr, nl, en in ws.iter_rows(min_row=2, values_only=True):
        if i is None:
            continue
        r = cur.execute('UPDATE artworks SET keywords_fr=?, keywords_nl=?, keywords_en=? WHERE id=?',
                        (fr, nl, en, int(i)))
        if r.rowcount:
            n += 1
        else:
            miss += 1
    report.append(f'keywords i18n : {n} artworks mis à jour, {miss} ids sans correspondance')

    # ── 5. ré-import des traductions corrigées ──────────────────────────
    # 5a. descriptions
    ws = wb['Descriptions']
    n = 0
    for i, _src, fr, nl, en in ws.iter_rows(min_row=2, values_only=True):
        if i is None:
            continue
        n += cur.execute('UPDATE artworks SET description_fr=?, description_nl=?, description_en=? WHERE id=?',
                         (fr, nl, en, int(i))).rowcount
    report.append(f'descriptions ré-importées : {n}')

    # 5b. questions
    ws = wb['Questions']
    n = 0
    for i, _src, fr, nl, en in ws.iter_rows(min_row=2, values_only=True):
        if i is None:
            continue
        n += cur.execute('UPDATE questions SET content_fr=?, content_nl=?, content_en=? WHERE id=?',
                         (fr, nl, en, int(i))).rowcount
    report.append(f'questions ré-importées : {n}')

    # 5c. types d'objets
    ws = wb["Types d'objets"]
    n = 0
    for i, _src, fr, nl, en in ws.iter_rows(min_row=2, values_only=True):
        if i is None:
            continue
        n += cur.execute('UPDATE types_of_object SET name_fr=?, name_nl=?, name_en=? WHERE id=?',
                         (fr, nl, en, int(i))).rowcount
    report.append(f"types d'objets ré-importés : {n}")

    # 5d. transcriptions (translation_*) — clé = artwork_id
    ws = wb['Transcriptions']
    n = 0
    for i, _src, fr, nl, en in ws.iter_rows(min_row=2, values_only=True):
        if i is None:
            continue
        n += cur.execute('''UPDATE transcriptions SET translation_fr=?, translation_nl=?, translation_en=?
                            WHERE artwork_id=?''', (fr, nl, en, int(i))).rowcount
    report.append(f'transcriptions (translation) ré-importées : {n}')

    # 5e. explications (explanation_*) — fr = source
    ws = wb['Transcriptions Explications']
    n = 0
    for i, _src, fr, nl, en in ws.iter_rows(min_row=2, values_only=True):
        if i is None:
            continue
        n += cur.execute('''UPDATE transcriptions SET explanation_fr=?, explanation_nl=?, explanation_en=?
                            WHERE artwork_id=?''', (fr, nl, en, int(i))).rowcount
    report.append(f'transcriptions (explanation) ré-importées : {n}')

    # 5f. témoignages enregistrés (interviews) — via RECORDED_IDS.
    # V8 : les interviews vivent dans `testimonies` avec visitor_id IS NULL et
    # leurs ids historiques ; RECORDED_IDS reste donc valable tel quel.
    # La feuille garde fichier_source (les colonnes speaker/source_file ont été
    # supprimées de la base en V3.1).
    ws = wb['Témoignages']
    n = miss = 0
    for q, _ville, date, _sl, fr, nl, en, sf in ws.iter_rows(min_row=2, values_only=True):
        if not sf:
            continue
        m = re.search(r'(\d+)', str(q or ''))
        qid = int(m.group(1)) if m else None
        rid = RECORDED_IDS.get(str(sf).strip(), {}).get(qid)
        created = date.date().isoformat() if hasattr(date, 'date') else (str(date)[:10] if date else None)
        content = {'fr': fr, 'nl': nl, 'en': en}.get(str(_sl or 'fr').strip().lower(), fr)
        r = cur.execute('''UPDATE testimonies
                           SET content=?, content_fr=?, content_nl=?, content_en=?, created_at=?
                           WHERE id=? AND visitor_id IS NULL''',
                        (content, fr, nl, en, created, rid))
        if rid is not None and r.rowcount == 1:
            n += 1
        else:
            miss += 1
    report.append(f'témoignages ré-importés : {n}, sans correspondance exacte : {miss}')

    # ── 6. retraits validés par le musée (relecture juillet 2026) ───────
    # 6a. 5 témoignages retirés du classeur (retrait volontaire confirmé) :
    #     Bernard_Kortrijk_out Q3, Claire_Kortrijk Q7, Godelieve_Kortrijk_out Q2/Q3/Q6
    removed = cur.execute(
        'DELETE FROM testimonies WHERE id IN (108, 112, 114, 127, 131) AND visitor_id IS NULL').rowcount
    report.append(f'témoignages retirés (validation musée) : {removed} supprimés')
    # 6b. type d'objet 28 « Printed image » : doublon du type 21 « Print - Image » ;
    #     reclassement de l'œuvre concernée puis suppression de l'entrée dupliquée
    moved = cur.execute(
        'UPDATE artworks SET type_of_object_id = 21 WHERE type_of_object_id = 28').rowcount
    dropped = cur.execute('DELETE FROM types_of_object WHERE id = 28').rowcount
    report.append(f'type 28 (doublon Printed image) : {moved} œuvre reclassée vers 21, {dropped} entrée supprimée')
    # 6c. nettoyage d'une clé coquillée historique du référentiel émotions
    cur.execute("DELETE FROM emotions WHERE emotion = 'excstacy'")

    con.commit()
    con.close()
    print('\n'.join('  • ' + r for r in report))


if __name__ == '__main__':
    main()
