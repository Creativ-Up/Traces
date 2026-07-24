#!/usr/bin/env python3
"""
Migration V9 de pp1_collection.db (PP1/Photomaton)
Colonnes Title/Author ajoutées par le musée dans PP1-Collection_Database.xlsx :
- ajoute artworks.author (TEXT, nullable — 14 œuvres attribuées, le reste anonyme)
  et l'importe depuis l'Excel musée ;
- les titres (title + title_fr/nl/en, colonnes créées en V4) sont importés par
  migrate_v3.py via la feuille « Titres » du classeur de relecture — lancer
  migrate_v3.py après cette migration.

Usage : python migrate_v9.py <db_path> <PP1-Collection_Database.xlsx> — idempotent.
"""
import sys
import sqlite3
from openpyxl import load_workbook

DB = sys.argv[1] if len(sys.argv) > 1 else 'pp1_collection.db'
XLSX = sys.argv[2] if len(sys.argv) > 2 else 'PP1-Collection_Database.xlsx'


def main():
    con = sqlite3.connect(DB)
    cur = con.cursor()
    report = []

    cols = [r[1] for r in cur.execute('PRAGMA table_info(artworks)')]
    if 'author' not in cols:
        cur.execute('ALTER TABLE artworks ADD COLUMN author TEXT')
        report.append('artworks.author ajoutée')
    else:
        report.append('artworks.author déjà présente')

    ws = load_workbook(XLSX, data_only=True)['Database']
    # en-têtes en ligne 2 ; Author = colonne 7 depuis l'ajout Title/Author
    assert str(ws.cell(2, 7).value).strip().lower() == 'author', 'colonne Author introuvable (structure Excel modifiée ?)'
    n = 0
    for r in range(3, ws.max_row + 1):
        i, a = ws.cell(r, 1).value, ws.cell(r, 7).value
        if i is None:
            continue
        a = str(a).strip() if a is not None and str(a).strip() else None
        n += cur.execute('UPDATE artworks SET author=? WHERE id=?', (a, int(i))).rowcount
    filled = cur.execute('SELECT COUNT(author) FROM artworks').fetchone()[0]
    report.append(f'auteurs importés : {n} œuvres traitées, {filled} avec auteur')

    con.commit()
    con.close()
    print('\n'.join('  • ' + r for r in report))


if __name__ == '__main__':
    main()
